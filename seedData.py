"""
-generate fake insurance data for testing
-includes intentional dirty data to practice cleaning

"""

import sqlite3
import random
from datetime import datetime, timedelta
from faker import Faker
from openpyxl import load_workbook

fake = Faker()
random.seed(42)
Faker.seed(42)
DB_PATH = "prototype.db"
NAICS_XLSX_PATH = "NAICsSICs.xlsx" 
SICS_XLSX_PATH = "SIC_CODES_FAMILIES.xlsx"

def load_NAICS_SIC_ROWS(path = NAICS_XLSX_PATH):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]
    wb.close()
    return rows 

def build_NAICS_codes(crosswalk_rows):
    seen = {}
    for naics_code, naics_description,_, _ in crosswalk_rows:
        seen.setdefault(str(naics_code), naics_description) 
    return list(seen.items())

def build_NAICS_SIC_crosswalk(crosswalk_rows):
    return [(str(n), str(s), sd) for n, s, sd in crosswalk_rows]


def make_dirty(value):
    choice = random.random()
    if choice < 0.15:
        return None
    elif choice < 0.30:
        return value.upper()
    elif choice < 0.45:
        return value.lower()
    elif choice < 0.55:
        return value + " "
    elif choice < 0.65:
        return value.replace("'", "")
    elif choice < 0.75:
        return value.replace(" ", "  ")
    else:
        return value
    
LOBS = [
    ("Auto-Commercial",),
    ("BOP",),
    ("Builders Risk",),
    ("Comprehensive Personal Liability",),
    ("Cyber",),
    ("Directors and Officers",),
    ("Dwelling Fire",),
     ("Employment Practices Liability",),
    ("Excess Casualty",),
    ("Excess Flood",),
    ("Excess Liability",),
    ("FNP Auto",),
    ("FNP E&S Excess",),
    ("FNP E&S Package",),
    ("General Liability",),
    ("General Liability - Discontinued Operations",),
    ("Habitational GL",),
    ("Mobile Home",),
    ("Occupational Accident",),
    ("Package",),
    ("Package Policy",),
    ("Primary Flood - Residential",),
    ("Professional Liability",),
    ("Property",),
    ("Religious Institutions Auto",),
    ("Religious Institutions E&S GL",),
    ("Religious Institutions E&S Package",),
    ("Religious Institutions E&S PROP",),
    ("Religious Institutions Package",),
    ("Renters",),
    ("Sexual Molestation Liability",),
    ("TRIA",),
    ("Umbrella",),
    ("Wind/Hail Deductible Buy Back",),
    ("Worker's Compensation",)
]

LOB_DIVISION_MAP = {
    "Auto-Commercial": ["Religious", "Non-Profit Package" , "Wholesale", "Other"],
    "BOP": ["Religious", "Wholesale", "Other"],
    "Builders Risk": ["Wholesale"],
    "Comprehensive Personal Liability": ["Other"],
    "Cyber": ["Non-Profit Package", "Wholesale", "Other"],
    "Directors and Officers": ["Religious", "Non-Profit Package", "Wholesale", "Other"],
    "Dwelling Fire": ["Other"],
    "Employment Practices Liability": ["Wholesale", "Other"],
    "Excess Casualty": ["Wholesale", "Other"],
    "Excess Flood": ["Non-Profit Package"],
    "Excess Liability": ["Other"],
    "FNP Auto": ["Non-Profit Package"],
    "FNP E&S Excess": ["Non-Profit Package"],
    "FNP E&S Package": ["Non-Profit Package"],
    "General Liability": ["Religious", "Non-Profit Package", "Wholesale", "Other"],
    "General Liability - Discontinued Operations": ["Wholesale"],
    "Habitational GL": ["Habitational", "Wholesale", "Other"],
    "Mobile Home": ["Other"],
    "Occupational Accident": ["Wholesale"],
    "Package": ["Religious", "Non-Profit Package", "Habitational", "Wholesale", "Other"],
    "Package Policy": ["Non-Profit Package", "Wholesale", "Other"],
    "Primary Flood - Residential": ["Non-Profit Package"],
    "Professional Liability": ["Wholesale", "Other"],
    "Property": ["Religious", "Non-Profit Package", "Wholesale", "Other"],
    "Religious Institutions Auto": ["Religious", "Non-Profit Package", "Other"],
    "Religious Institutions E&S GL": ["Religious"],
    "Religious Institutions E&S Package": ["Religious"],
    "Religious Institutions E&S PROP": ["Religious"],
    "Religious Institutions Package": ["Religious", "Other"],
    "Renters": ["Other"],
    "Sexual Molestation Liability": ["Wholesale", "Other"],
    "TRIA": ["Wholesale"],
    "Umbrella": ["Religious", "Non-Profit Package", "Habitational", "Wholesale", "Other"],
    "Wind/Hail Deductible Buy Back": ["Wholesale", "Other"],
    "Worker's Compensation": ["Religious", "Non-Profit Package", "Wholesale", "Other", "Workers Compensation"],
}

"""
LOBS = [
    ("Auto-Commercial",),
    ("BOP",),
    ("Builders Risk",),
    ("Comprehensive Personal Liability",),
    ("Cyber",),
    ("Deductible Buy back",),
    ("Directors and Officers",),
    ("Dwelling Fire",),
    ("Earthquake",),
    ("Employment Practices Liability",),
    ("Excess Casualty",),
    ("Excess Comprehensive Personal Liability",),
    ("Excess Flood",),
    ("Excess Liability",),
    ("FL Homeowners with Wind",),
    ("FNP Auto",),
    ("FNP E&S Excess",),
    ("FNP E&S Package",),
    ("General Liability - Discontinued Operations",),
    ("General Liability",),
    ("Habitational GL",),
    ("Homeowners",),
    ("Inland Marine",),
    ("Mobile Home",),
    ("Motor Home",),
    ("Motorcycle",),
    ("Package",),
    ("Package Policy",),
    ("Personal Articles Floater",),
    ("Personal Auto",),
    ("Personal Excess Umbrella",),
    ("Personal Umbrella",),
    ("Pollution Liability",),
    ("Premier Home",),
    ("Primary Flood - Commercial",),
    ("Primary Flood - Residential",),
    ("Professional Liability",),
    ("Property",),
    ("Religious Institutions Auto",),
    ("Religious Institutions E&S GL",),
    ("Religious Institutions E&S Package",),
    ("Religious Institutions E&S PROP",),
    ("Religious Institutions Package",),
    ("Renters",),
    ("Sexual Molestation Liability",),
    ("TRIA",),
    ("Umbrella",),
    ("Wind Only",),
    ("Wind/Hail Deductible Buy Back",),
    ("Worker's Compensation",),
]
"""
CARRIERS = [
    ("CAR001", "A.I.M. Mutual Insurance Co."),
    ("CAR002", "Admiral Insurance Company"),
    ("CAR003", "AIG Property Casualty Company"),
    ("CAR004", "Alpha Property & Casualty Insurance Co"),
    ("CAR005", "American Bankers Insurance Company of FL"),
    ("CAR006", "AmGUARD Insurance Company"),
    ("CAR007", "AmTrust Insurance Co of Kansas"),
    ("CAR008", "AmTrust North America Inc."),
    ("CAR009", "Arch Insurance Company"),
    ("CAR010", "ARI Insurance Company"),
    ("CAR011", "Aspen American Insurance Company"),
    ("CAR012", "Aspen Specialty Insurance Company"),
    ("CAR013", "Associated Industries Insurance Company"),
    ("CAR014", "Axis Insurance Company"),
    ("CAR015", "AXIS Surplus Insurance Company"),
    ("CAR016", "Benchmark Insurance Company"),
    ("CAR017", "Berkley National Insurance Company"),
    ("CAR018", "Berkley National Insurance Company - Senior Living Admitted"),
    ("CAR019", "Berkley Specialty Insurance Company - Senior Living E & S"),
    ("CAR020", "Berkley Specialty Insurance Company -- E & S Non-Profit"),
    ("CAR021", "Blue Ridge Indemnity Insurance"),
    ("CAR022", "BMS Group"),
    ("CAR023", "Boost.io"),
    ("CAR024", "Bristol West Insurance Company"),
    ("CAR025", "Business Risk Partners"),
    ("CAR026", "Capitol Specialty Insurance Corporation"),
    ("CAR027", "Carolina Casualty Insurance Company"),
    ("CAR028", "Century Surety Co"),
    ("CAR029", "Certain Underwriters at Lloyd's -Coalition"),
    ("CAR030", "Certain Underwriters at Lloyds"),
    ("CAR031", "Charter Indemnity Co"),
    ("CAR032", "Chubb Indemnity Insurance Company"),
    ("CAR033", "Chubb Insurance Company of New Jersey"),
    ("CAR034", "Church Mutual Insurance Company - Non-Profit Program"),
    ("CAR035", "Church Mutual Insurance Company - Religious COE"),
    ("CAR036", "Clearance Carrier"),
    ("CAR037", "CM Regent Insurance Company"),
    ("CAR038", "CM Vantage Specialty Insurance Company"),
    ("CAR039", "CNA"),
    ("CAR040", "CO/ACTION Specialty Insurance Group"),
    ("CAR041", "Coast National Insurance Company"),
    ("CAR042", "Convex Insurance UK Limited"),
    ("CAR043", "Dairyland County Mutual Insurance Company of Texas"),
    ("CAR044", "Dairyland Insurance Company"),
    ("CAR045", "Drive New Jersey Insurance Company"),
    ("CAR046", "Employers Assurance Co - Healthcare"),
    ("CAR047", "Employers Assurance Co - Non-Profit"),
    ("CAR048", "Employers Assurance Company"),
    ("CAR049", "Employers Compensation Ins Co - Healthcare"),
    ("CAR050", "Employers Compensation Ins Co - Non-Profit"),
    ("CAR051", "Employers Preferred Insurance Company"),
    ("CAR052", "Endurance American Specialty Insurance Company"),
    ("CAR053", "Evanston Insurance Company"),
    ("CAR054", "Federal Insurance Company"),
    ("CAR055", "Financial Indemnity Company"),
    ("CAR056", "Foremost County Mutual Insurance Company"),
    ("CAR057", "Foremost Insurance Company"),
    ("CAR058", "Foremost Lloyds of TX"),
    ("CAR059", "Fortegra Specialty Insurance Company"),
    ("CAR060", "General Star Indemnity Company"),
    ("CAR061", "Great American Insurance Company"),
    ("CAR062", "Guard"),
    ("CAR063", "Guard Insurance Company"),
    ("CAR064", "GuideOne Elite Insurance Company"),
    ("CAR065", "GuideOne Mutual Insurance Company"),
    ("CAR066", "GuideOne National Insurance Company"),
    ("CAR067", "GuideOne Specialty Insurance Company"),
    ("CAR068", "Hadron Specialty Insurance Company"),
    ("CAR069", "Harleysville Insurance Company"),
    ("CAR070", "Harleysville Worcester Ins Co"),
    ("CAR071", "Hartford Casualty Ins Co"),
    ("CAR072", "Highview National Insurance Company"),
    ("CAR073", "Houston Casualty Company"),
    ("CAR074", "Hudson Excess Ins Co - Corvus"),
    ("CAR075", "Hudson Excess Insurance Company"),
    ("CAR076", "Hudson Insurance Company"),
    ("CAR077", "Hudson Specialty Insurance Company"),
    ("CAR078", "Insurance Company of the South"),
    ("CAR079", "Insurmark/Floodwatch"),
    ("CAR080", "Ironshore Europe Limited/Certain Underwriters at Lloyds"),
    ("CAR081", "Ironshore Specialty Insurance Company"),
    ("CAR082", "Kinsale Insurance Company"),
    ("CAR083", "Lexington Insurance Co"),
    ("CAR084", "Lexington Insurance Company"),
    ("CAR085", "LIO Insurance Group"),
    ("CAR086", "Lloyd's of London - Beazley Syndicate #2323/0623"),
    ("CAR087", "Lloyds of London"),
    ("CAR088", "Lyndon Southern Insurance Company"),
    ("CAR089", "Markel Ins Company"),
    ("CAR090", "Markel Insurance Company"),
    ("CAR091", "Middlesex Insurance Company"),
    ("CAR092", "Motor Transport Mutual RRG"),
    ("CAR093", "Moxie Series (Admitted)"),
    ("CAR094", "National Continental Insurance Company"),
    ("CAR095", "National General Insurance Company"),
    ("CAR096", "North American Capacity Ins Co"),
    ("CAR097", "North American Specialty Insurance Corporation"),
    ("CAR098", "Pacific Indemnity Company"),
    ("CAR099", "Palms Insurance Company, Ltd"),
    ("CAR100", "Peak Property and Casualty Insurance Corporation"),
    ("CAR101", "Pennsylvania Manufacturers Assoc. Insurance Co"), 
    ("CAR102", "Privilege Underwriters Reciprocal Exchange"),
    ("CAR103", "Progressive Casualty Insurance Company"),
    ("CAR104", "Progressive County Mutual Insurance Co"),
    ("CAR105", "PURE Programs, LLC"), 
    ("CAR106", "Quantum Series"), 
    ("CAR107", "Reliable Lloyds Insurance Company"),
    ("CAR108", "Republic-Vanguard Insurance Company"),
    ("CAR109", "Risk Smith"), 
    ("CAR110", "Rochdale Ins Co Of NY"),
    ("CAR111", "Security National Insurance Co"), 
    ("CAR112", "Sentry Insurance a Mutual Company"), 
    ("CAR113", "Sequoia Ins Co"),
    ("CAR114", "Sompo"),
    ("CAR115", "StarStone National Insurance Company"),
    ("CAR116", "Surya Ins Company - NEMT"),
    ("CAR117", "Technology Ins Co Inc"),
    ("CAR118", "Templar Series"),
    ("CAR119", "The Hartford"),
    ("CAR120", "The Hartford - VCAP"),
    ("CAR121", "Travelers Excess and Surplus LInes Company"),
    ("CAR122", "Unitrin County Mutual Insurance Co"),
    ("CAR123", "Vigilant Insurance Company"),
    ("CAR124", "Viking Insurance Company of Wisconsin"),
    ("CAR125", "Wesco Ins Co"),
    ("CAR126", "Westchester, A Chubb Company"),
    ("CAR127", "Wyvern Series (Admitted)"),
    ("CAR128", "Wyvern Series (Non-Admitted)")

]
"""
(broker_id, producer_name, producer_state, relationship_owner)
found 907 distinct producers so I'm generating 200 fake ones 
"""
BROKER_GROUP_ID = 'BRKGRP001'  #global var to be accessed by numerous fcts
BROKER_GROUP_NAME = "Convelo"
BROKER_GROUPS = [
    (BROKER_GROUP_ID,
     BROKER_GROUP_NAME)
]

# found over 5000 distinct account names so generating fake ones 



TYPES = [
    "New",
    "Broker on Record (BOR)",
    "Courtesy Filing",
    "Purchased Book of Business",
    "Renewal",
    "Rewrite",
    "Roll-Over"
]

PRIORITIES = [
    "Null",
    "0 - No opportunity, log in only ?Does not fit appetite",
    "1 - Received application and some loss data, account not qualified",
    "2 - Received application, account qualified ? Additional info requested to complete file",
    "3 - Fully qualified submission with all required information obtained from the broker-No commitment from Broker",
    "4 - Fully complete submission with broker requirements to bind and will be recommended -Broker New opportunity",
    "5 - Fully complete submission with broker requirements to bind and will be recommended ? Broker Controls"
]

SUBMISSION_STATUSES = [
    "Additional Information Received",
    "Additional Information Request",
    "Bound",
    "Bound-Issued",
    "Cancelled",
    "Cancelled by Correction",
    "Declined",
    "Expired",
    "In Rating",
    "Indicated",
    "Lost",
    "Lost on BOR",
    "Non-Renewed",
    "Not Quoted",
    "Not Taken Up",
    "Pending Cancellation",
    "Pending Reinstatement",
    "Quoted",
    "Quoted not Bound",
    "Referred to Carrier",
    "Submitted",
    "Unbound Correction",
    "Unbound Endorsement",
    "Unbound Internal Correction",
    "Underwriting Review",
    "Unknown Status"
]

QUOTE_STATUSES = [
    "Quoted",
    "Quoted not Bound",
    "Bound-Issued",
    "Bound",
    "Declined",
    "Not Taken Up",
    "Lost"
]

POLICY_STATUSES = [
    "Bound",
    "Bound-Issued",
    "Expired",
    "Cancelled",
    "Pending Cancellation"
]

LOSS_STATUSES = [
    "open",
    "closed",
    "closed without payment",
    "reopened",
    "pending"
]

DIVISIONS = [
    "Wholesale",
    "Habitational",
    "Non-Profit Package",
    "Religious",
    "Workers Compensation",
    "Other"
]

def build_accounts(n=500, naics_pool = None):
    rows = []
    for i in range(1, n+1):
        account_name = fake.company()
        employee_count = random.randint(1, 5000)
        annual_revenue = round(random.uniform(50000, 50000000), 2)
        naics_code = random.choice(naics_pool)
        insured_state = fake.state.abbr()
        rows.append((f"ACC{str(i).zfill(3)}",
            make_dirty(account_name),  #dirty
            account_name,              #clean
            employee_count,
            annual_revenue,
            naics_code, 
            insured_state
   ))
    return rows

n = 100
def build_brokers(n, group_id = BROKER_GROUP_ID):
    rows = []
    for i in range(1, n+1):
        broker_name = fake.company() + random.choice([" Insurance", " Risk Partners", " & Associates", " Group", ""])
        relationship_owner = fake.name()
        producer_state = fake.state_abbr()
        rows.append ((
            (f"BRK{str(i).zfill(3)}"),
            broker_name,
            group_id,
            make_dirty(relationship_owner),
            make_dirty(producer_state)
        ))
    return rows

n = 400
def build_producers(n, brokers): 
    rows = []
    firm_ids = [b[0] for b in brokers]

    n_large = int(len(firm_ids)*0.15)  #0.15 represents 15% of firms being large and 85% of firms being small.
    n_small = len(firm_ids) - n_large
    large_firms = firm_ids[:n_large]
    small_firms = firm_ids[n_large:]

    SMALL_FIRM_SIZE = 2
    LARGE_FIRM_SIZE = (n - n_small * SMALL_FIRM_SIZE) // n_large
    i = 1

    for firm_id in large_firms:
        for _ in range (LARGE_FIRM_SIZE):
            rows.append((
                f"PRD{str(i).zfill(3)}",
                fake.name(),
                random.choice(['R','W', 'MGA']),
                firm_id
            ))
            i+=1

    for firm_id in small_firms:
        for _ in range(SMALL_FIRM_SIZE):
            rows.append((
                f"PRD{str(i).zfill(3)}",
                fake.name(),
                random.choice(['R','W', 'MGA']),
                firm_id
            ))
            i+=1
    return rows

def build_submission_groups(accounts, producers):
    rows = []
    for account in accounts:
        
       # 1. randomly picks how many submission groups this account will have -> 1, 2, 3
       # 2. loops based on # of sub_groups
       # 3. generate group submision number 
       # 4. generates a random decimal number 0-20,000,000 for target_premium
        
        # how many groups per account
        n_groups = random.randint(1,3)
        for i in range(n_groups):
            rows.append((
               (f"GRP{str(len(rows)+1).zfill(4)}"),
               round(random.uniform(0, 20000000), 2),
               account[0],
               random.choice(producers)[0]
            ))

    return rows
    
def build_submissions(submission_groups, brokers):
    rows = []
    today = datetime.today()
    start = today.replace(month=1, day=1)
    end = today.replace(month=12, day=31)
    for i,  group in enumerate(submission_groups):
            # how many submissions per group; same pattern as n_groups just one level deeper 
        n_submissions = random.randint(1,5) 
        for j in range(n_submissions):
            eff = fake.date_between(start_date=start, end_date=end)
            exp = eff + timedelta(days=365)
            lob_name = random.choice(LOBS)[0]
            rows.append((
                (f"CON{str(i).zfill(3)}{str(j).zfill(2)}"),
                group[2],
                random.choice(brokers)[0],
                lob_name,
                eff.strftime("%Y-%m-%d"),
                exp.strftime("%Y-%m-%d"),
                round(random.uniform(0, 20000000), 2),
                random.choice(PRIORITIES),
                random.choice(TYPES),
                group[0],
                random.choice(SUBMISSION_STATUSES),
                random.choice(LOB_DIVISION_MAP.get(lob_name, DIVISIONS))
                ))
            
    return rows
        
def build_quotes(submissions):
        rows = []
        for i, submission in enumerate(submissions):
            n_quotes = random.randint(1,3)
            for j in range(n_quotes):
                rows.append((
                    (f"QUO{str(i).zfill(3)}{str(j).zfill(2)}"),
                    submission[0],
                    random.choice(CARRIERS)[0],
                    fake.name(),
                    round(random.uniform(0, 20000000), 2),
                    random.choice(QUOTE_STATUSES)
                ))    
        return rows 
            
def build_policies(quotes):
        rows = []
        for i, quote in enumerate(quotes):
            if quote[5] in ("Bound", "Bound-Issued"):
                rows.append((
                    (f"POL{str(i).zfill(3)}"),
                    quote[0],
                    quote[2],
                    round(random.uniform(5000, 500000), 2),
                    random.choices(POLICY_STATUSES, weights = [25, 28, 20, 15, 12], k=1)[0]
                ))
        return rows        
    
def build_losses(policies):
        rows = []
        for i, policy in enumerate(policies):
            if random.random() < 0.25: 
               loss_date = fake.date_between(
                   start_date = datetime(2024, 1, 1),
                   end_date = datetime.today())
               report_date = fake.date_between(
                   start_date = loss_date,
                   end_date = datetime.today())
               paid_amount = round(random.uniform(0, 500000), 2)
               reserved_amount = round(random.uniform(0, 500000), 2)
               incurred_amount = round(paid_amount + reserved_amount, 2)
               rows.append((
                    (f"LOS{str(i).zfill(3)}"),
                    policy[0],
                    loss_date.strftime("%Y-%m-%d"),
                    report_date.strftime("%Y-%m-%d"),
                    paid_amount,
                    reserved_amount,
                    incurred_amount,
                    random.choice(LOSS_STATUSES)
                ))
        return rows
    
def seed():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    with open("prototypeSchema.sql") as f:
        conn.executescript(f.read())

    crosswalk_rows = load_NAICS_SIC_ROWS()
    naics_codes = build_NAICS_codes(crosswalk_rows)
    naics_sic = build_NAICS_SIC_crosswalk(crosswalk_rows)
    cur.executemany("INSERT OR REPLACE INTO NAICS VALUES (?,?)", naics_codes)
    cur.executemany("INSERT OR REPLACE INTO NAICS_SIC_CROSSWALK VALUES (?,?,?)", naics_sic)
    cur.executemany("INSERT OR REPLACE INTO lines_of_business VALUES (?)", LOBS)
    cur.executemany("INSERT OR REPLACE INTO broker_groups VALUES (?,?)", BROKER_GROUPS)
    carriers_with_underwriters = [
        (carrier_id, carrier_name, make_dirty(fake.name())) for carrier_id, carrier_name in CARRIERS
    ]
    cur.executemany("INSERT OR REPLACE INTO carriers VALUES (?,?,?)", carriers_with_underwriters)
    
    naics_pool = [c[0] for c in naics_codes]
    
    accounts = build_accounts(500, naics_pool)
    cur.executemany("INSERT OR REPLACE INTO accounts VALUES (?,?,?,?)", accounts)

        # transactional data in dependency order 
        # each build_ fct receives its parent as a parameter
        # ? must match the number of fields in data schema per table
    brokers = build_brokers(n, group_id = BROKER_GROUP_ID)
    cur.executemany("INSERT OR REPLACE INTO brokers VALUES (?,?,?,?,?)", brokers)

    producers = build_producers(n, brokers)
    cur.executemany("INSERT OR REPLACE INTO producers VALUES (?,?,?,?)", producers)

    submission_groups = build_submission_groups(accounts, producers)
    cur.executemany("INSERT OR REPLACE INTO submission_groups VALUES (?,?,?,?)", submission_groups)

    submissions = build_submissions(submission_groups, brokers)
    cur.executemany("INSERT OR REPLACE INTO submissions VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", submissions)

    quotes = build_quotes(submissions)
    cur.executemany("INSERT OR REPLACE INTO quotes VALUES (?,?,?,?,?,?)", quotes)

    policies = build_policies(quotes)
    cur.executemany("INSERT OR REPLACE INTO policies VALUES (?,?,?,?,?)", policies)

    losses = build_losses(policies)
    cur.executemany("INSERT OR REPLACE INTO losses VALUES (?,?,?,?,?,?,?,?)", losses)

    conn.commit()
    conn.close()

if __name__ == "__main__":
    seed()
